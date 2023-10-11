import openpyxl
import pathlib
import csv
from tkinter import filedialog

target_folder_name = filedialog.askdirectory(
    title="CSVファイルが格納されているフォルダを選択してください。", initialdir="./"
)
target_folder = pathlib.Path(target_folder_name)

new_book = openpyxl.Workbook()

for target_csv in target_folder.glob("*.csv"):
    sheet = new_book.create_sheet(title=target_csv.name)

    csv_file = open(target_csv, encoding="UTF-8")
    reader = csv.reader(csv_file)

    for i, row in enumerate(reader, start=1):
        for j, data in enumerate(row, start=1):
            sheet.cell(i, j).value = data
    print(f"{target_csv.name}をExcelファイルに転記しました。")

if len(new_book.sheetnames) >= 2:
    del new_book["Sheet"]

new_book.save("assembled_csv.xlsx")
input("Enterキーでプログラムを閉じます：")
